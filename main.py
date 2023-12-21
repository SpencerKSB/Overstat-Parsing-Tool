import csv
import os
import sys
import openpyxl
from pathlib import Path
from random import randint

from PyQt5 import QtWidgets
from PyQt5.QtCore import QObject, QThread, pyqtSignal
from PyQt5.QtWidgets import *
from PyQt5.QtCore import QTimer

StyleSheet = '''
    #BlueProgressBar {
        min-height: 12px;
        max-height: 12px;
        border-radius: 6px;
    }
    #BlueProgressBar::chunk {
        border-radius: 6px;
        background-color: #0f76d6;
    }
    '''

RawDataFile = ()
FightsFile = ()


class Main(QObject):

    finished = pyqtSignal(str)
    progress = pyqtSignal(int)

    def xlsx_append_csv(self, wb, file):
        sheet = wb.create_sheet(f"data")
        with open(file, newline='', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile)
            line_num = 0
            for row in reader:
                self.write_row(row, sheet, line_num+1)
                line_num += 1

        return wb

    def csv_split(self, file):
        l = file.split("/")
        name = l[len(l)-1].split(".")[0]
        maps = {}
        with open(file, newline='', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile)

            header = next(reader)
            for row in reader:
                if not row[0] in maps.keys():
                    maps[row[0]] = [header]
                maps[row[0]].append(row)

        workbook = openpyxl.Workbook()
        count = 1
        for m in maps:
            sheet = workbook.create_sheet(f"Map {count}")

            line_num = 0
            for line in maps[m]:
                self.write_row(line, sheet, line_num+1)
                line_num += 1

            count += 1

        file_name = get_path() + f"\\Parsed Scrim Log - {name}.xlsx"
        return [workbook, file_name]

    def write_row(self, row, sheet, row_num, start_col=1):
        for column_num, val in enumerate(row, start_col):
            if is_number(val):
                sheet.cell(row=row_num, column=column_num).value = float(val)
            else:
                sheet.cell(row=row_num, column=column_num).value = val

    def run(self):
        try:
            global RawDataFile
            global FightsFile

            data = FightsFile[0]  # Fights with Interface
            data2 = RawDataFile[0]  # Raw Data

            arr = self.csv_split(data2)
            print("Successfully split data2")
            name = arr[1]
            wb = arr[0]

            final_xls = self.xlsx_append_csv(wb, data)
            print("Successfully appended data")

            final_xls.save(name)
            print(f"Successfully saved {name}")

            self.finished.emit(f"Successfully exported file: {name}")
        except Exception as e:
            self.finished.emit(f'Unknown Error: {e}')


class GUI(QtWidgets.QMainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setupUi()

    def setupUi(self):
        self.setWindowTitle("OTC Stats Tool")
        self.resize(500, 200)
        self.centralWidget = QtWidgets.QWidget()
        self.setCentralWidget(self.centralWidget)

        # Create and connect widgets
        self.runSplitBtn = QtWidgets.QPushButton("Split", self)
        self.runSplitBtn.clicked.connect(self.runSplit)
        self.progressBar = ProgressBar(self, minimum=0, maximum=0, textVisible=False, objectName="BlueProgressBar")
        self.progressBar.setVisible(False)


        # Raw Data - data2
        rawDataLabel = QtWidgets.QLabel("Raw Data CSV: ")
        self.rawDataTextBox = QtWidgets.QLineEdit()
        self.rawDataBtn = QtWidgets.QPushButton("Open", self)
        self.rawDataBtn.clicked.connect(self.getRawDataFile)

        rawDataContainer = QtWidgets.QHBoxLayout()
        rawDataContainer.addWidget(self.rawDataTextBox)
        rawDataContainer.addWidget(self.rawDataBtn)


        # Fights with Interface - data
        fightsLabel = QtWidgets.QLabel("Fights with Interface: ")
        self.fightsTextBox = QtWidgets.QLineEdit()
        self.fightsBtn = QtWidgets.QPushButton("Open", self)
        self.fightsBtn.clicked.connect(self.getFightsFile)

        fightsContainer = QtWidgets.QHBoxLayout()
        fightsContainer.addWidget(self.fightsTextBox)
        fightsContainer.addWidget(self.fightsBtn)


        # data container
        dataLayout = QtWidgets.QFormLayout()
        dataLayout.addRow(rawDataLabel, rawDataContainer)
        dataLayout.addRow(fightsLabel, fightsContainer)

        # Main layout
        layout = QtWidgets.QVBoxLayout()
        layout.addLayout(dataLayout)
        layout.addWidget(self.runSplitBtn)
        layout.addWidget(self.progressBar)
        self.centralWidget.setLayout(layout)

    def getRawDataFile(self):
        file_filter = 'Data File (*.csv)'
        res = QFileDialog().getOpenFileName(
            parent=self,
            caption='Select the Raw Data File',
            directory=os.getcwd(),
            filter=file_filter,
            initialFilter='Data File (*.csv)'
        )

        global RawDataFile
        RawDataFile = res

        print(res)
        self.rawDataTextBox.setText(res[0])

    def getFightsFile(self):
        file_filter = 'Data File (*.csv)'
        res = QFileDialog().getOpenFileName(
            parent=self,
            caption='Select the Fights File',
            directory=os.getcwd(),
            filter=file_filter,
            initialFilter='Data File (*.csv)'
        )
        global FightsFile
        FightsFile = res
        print(res)
        self.fightsTextBox.setText(res[0])

    def runSplit(self):
        if len(RawDataFile) < 2 or len(FightsFile) < 2:
            self.dispMsg('Please Select a File')
            return

        f1 = Path(RawDataFile[0])
        f2 = Path(FightsFile[0])

        if not f1.is_file() or not f2.is_file():
            self.dispMsg('Selected item is not a file')
            return

        if RawDataFile[1] != 'Data File (*.csv)' or FightsFile[1] != 'Data File (*.csv)':
            self.dispMsg('Invalid File Type Selected. Ensure both files are CSVs')
            return

        self.progressBar.setVisible(True)
        self.runSplitBtn.setDisabled(True)
        self.thread = QThread()
        self.worker = Main()
        self.worker.gui_obj = self

        self.worker.moveToThread(self.thread)

        self.thread.started.connect(self.worker.run)
        self.worker.finished.connect(self.thread.quit)
        self.worker.finished.connect(self.worker.deleteLater)
        self.thread.finished.connect(self.thread.deleteLater)
        self.worker.finished.connect(self.job_completion)

        self.thread.start()

    def job_completion(self, txt):
        self.progressBar.setVisible(False)
        self.runSplitBtn.setDisabled(False)
        print(txt)
        self.dispMsg(txt)

    def dispMsg(self, txt):
        msg = QMessageBox()
        msg.setText(txt)
        msg.setWindowTitle("OTC Stats Tool")
        msg.setStandardButtons(QMessageBox.Ok)
        e = msg.exec_()


class ProgressBar(QProgressBar):

    def __init__(self, *args, **kwargs):
        super(ProgressBar, self).__init__(*args, **kwargs)
        self.setValue(0)
        if self.minimum() != self.maximum():
            self.timer = QTimer(self, timeout=self.onTimeout)
            self.timer.start(randint(1, 3) * 1000)

    def onTimeout(self):
        if self.value() >= 100:
            self.timer.stop()
            self.timer.deleteLater()
            del self.timer
            return
        self.setValue(self.value() + 1)


def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        return False


def get_path():
    return os.path.dirname(os.path.realpath(sys.argv[0]))


if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setStyleSheet(StyleSheet)
    win = GUI()
    win.show()
    sys.exit(app.exec())
